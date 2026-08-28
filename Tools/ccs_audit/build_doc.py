#!/usr/bin/env python3
"""Render the CCS audit workbook to HTML, and from there to PDF.

The three files -- .xlsx, .html, .pdf -- said different things on 2026-08-28
because the workbook was edited and the other two were not.  The spreadsheet
is the single source; this makes the other two fall out of it.

    python3 Tools/ccs_audit/build_doc.py            # html only
    python3 Tools/ccs_audit/build_doc.py --pdf      # and render with soffice

The header prose and the closing appendix are NOT in the spreadsheet -- they
are page furniture -- so they are carried over verbatim from the existing
HTML.  Everything between them is generated.
"""
import argparse
import html
import re
import subprocess
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parents[2]
DOC = HERE / "docs/job_model/CCS_State_Action_Result"

#: (spreadsheet sheet, printed heading, column widths).  The widths are the
#: hand-tuned ones from the first edition; they are what keeps a 6-column
#: landscape table readable, and nothing derives them.
SECTIONS = [
    ("1 AGV Task States", "1. AGV task states (§2.3, §4.6.6, §5.12)",
     [10, 13, 20, 15, 4, 18]),
    ("2 Post-task States", "2. Post-task states (§4.6.6)",
     [11, 15, 15, 15, 4, 18]),
    ("3 Cold Press to Diecut", "3. Cold press &rarr; diecut (§1.3)",
     [13, 13, 26, 15, 4, 19]),
    ("4 Diecut to Winding", "4. Diecut &rarr; winding (§1.2.1)",
     [12, 13, 27, 17, 5, 18]),
    ("5 Empty Bobbin Return", "5. Empty bobbin return (§1.2.2)",
     [12, 14, 25, 17, 4, 18]),
    ("6 Posting Inventory", "6. Posting and inventory (§4.6.2)",
     [13, 15, 18, 21, 4, 19]),
    ("7 State Vocabulary", "7. State vocabulary (§2.2, §5.3)", [14, 20, 55]),
    ("8 Parameters", "8. Parameters that gate the transitions", [22, 7, 60]),
    ("9 Open Questions", "9. Open questions", [10, 45, 42]),
    ("10 Implemented", "10. What we have built", [20, 9, 21, 47]),
    ("11 Not Implemented", "11. What we have not built", [19, 36, 44]),
    ("12 Live snapshot", None, [6, 7, 9, 20, 26, 22]),
]

#: Rendered AFTER the closing appendix, not before it.  The appendix ends on
#: the simulator screenshot, and this table is the legend for it: which layer
#: owns what, and how much of each already runs.
POST_APPENDIX = [
    ("13 Layers", "13. The layers, and what each one already does", [22, 78]),
]


def cell(v):
    """Escape, then honour the newlines the spreadsheet author typed.

    A cell holding a bulleted list is written with real newlines in the
    workbook.  HTML collapses those to spaces, so without this the bullets
    run together into one paragraph and the table becomes unreadable.
    """
    if v is None:
        return ""
    return html.escape(str(v)).replace("\n", "<br/>")


def table(ws, widths):
    cols = "".join(f'<col style="width:{w}%"/>' for w in widths)
    head = "".join(f"<th>{cell(c.value)}</th>" for c in ws[1][:len(widths)])
    body = []
    for row in ws.iter_rows(min_row=2):
        if all(c.value in (None, "") for c in row):
            continue
        body.append("<tr>" + "".join(f"<td>{cell(c.value)}</td>"
                                     for c in row[:len(widths)]) + "</tr>")
    return (f"<table>{cols}<thead><tr>{head}</tr></thead>"
            f"<tbody>{''.join(body)}</tbody></table>")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf", action="store_true",
                    help="render the HTML with libreoffice afterwards")
    args = ap.parse_args()

    old = (DOC.with_suffix(".html")).read_text(encoding="utf-8")
    header = old[:old.index("<h2")]
    # The appendix runs from its own heading to the first section that comes
    # AFTER it -- not to the end of the file.  Slicing to the end re-swallowed
    # the generated post-appendix sections, so every build appended another
    # copy of them and the file grew by one section each time.
    start = old.index("<h2 class='brk'>12. Appendix")
    end = len(old)
    for _, heading, _ in POST_APPENDIX:
        marker = f"<h2 class='brk'>{heading}</h2>"
        if marker in old:
            end = min(end, old.index(marker))
    appendix = old[start:end]
    # The old tail closed the document.  Sections now follow it, so the
    # closing tags are held back and re-added last -- otherwise section 13
    # renders outside <body> and LibreOffice silently drops it.
    closer = "</body></html>"
    for tag in ("</body></html>", "</body>", "</html>"):
        if appendix.rstrip().endswith(tag):
            appendix = appendix.rstrip()[:-len(tag)]
            break

    # The live-snapshot heading carries its own timestamp, so it is read back
    # rather than written out -- the snapshot is only as current as the run
    # that produced it, and pretending otherwise is how a stale table looks
    # authoritative.
    snap_heading = re.search(r"<h2 class='brk'>(12\. Live snapshot[^<]*)</h2>",
                             old).group(1)

    wb = openpyxl.load_workbook(DOC.with_suffix(".xlsx"))
    parts = [header]
    for sheet, heading, widths in SECTIONS:
        title = heading if heading is not None else snap_heading
        brk = "" if parts == [header] else " class='brk'"
        parts.append(f"<h2{brk}>{title}</h2>{table(wb[sheet], widths)}")
    parts.append(appendix)
    for sheet, heading, widths in POST_APPENDIX:
        parts.append(f"<h2 class='brk'>{heading}</h2>{table(wb[sheet], widths)}")

    parts.append(closer)
    out = "".join(parts)
    DOC.with_suffix(".html").write_text(out, encoding="utf-8")
    print(f"wrote {DOC.with_suffix('.html').relative_to(HERE)}"
          f"  ({len(out):,} bytes)")

    if args.pdf:
        # LibreOffice, because it is what produced the first edition and the
        # page furniture is tuned to its @page handling.
        subprocess.run(
            ["soffice", "--headless", "--convert-to", "pdf",
             "--outdir", str(DOC.parent), str(DOC.with_suffix(".html"))],
            check=True, stdout=subprocess.DEVNULL)
        print(f"wrote {DOC.with_suffix('.pdf').relative_to(HERE)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
